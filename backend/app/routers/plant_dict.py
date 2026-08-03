from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete as sa_delete
from typing import List, Optional
import io

from openpyxl import load_workbook
from pydantic import BaseModel

from app.database import get_db
from app.models.plant_dict import Plant, Line, Station
from app.models.inspection import Inspection
from app.models.exception import ExceptionTicket
from app.utils.response import success, error
from app.schemas.plant_dict import PlantTreeOut, LineOut, StationOut

router = APIRouter(prefix="/dict", tags=["厂区字典"])


class PlantUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None


class LineUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None


class StationUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None


class PlantCreate(BaseModel):
    code: str
    name: str


class LineCreate(BaseModel):
    plant_id: str
    code: str
    name: str


class StationCreate(BaseModel):
    line_id: str
    code: str
    name: str


@router.get("/tree")
async def get_plant_dict_tree(db: AsyncSession = Depends(get_db)):
    """
    获取厂区-线别-站别树形结构（使用 JOIN 优化）
    """
    # 使用 joinedload 一次性加载完整树结构
    from sqlalchemy.orm import joinedload

    # 查询所有厂区及其线别
    plant_result = await db.execute(
        select(Plant).options(
            joinedload(Plant.lines).joinedload(Line.stations)
        )
    )
    plants = plant_result.unique().scalars().all()

    tree = []
    for plant in plants:
        plant_data = {
            "id": str(plant.id),
            "code": plant.code,
            "name": plant.name,
            "children": []
        }
        for line in (plant.lines if hasattr(plant, 'lines') else []):
            line_data = {
                "id": str(line.id),
                "plant_id": str(line.plant_id),
                "code": line.code,
                "name": line.name,
                "children": []
            }
            for station in (line.stations if hasattr(line, 'stations') else []):
                line_data["children"].append({
                    "id": str(station.id),
                    "line_id": str(station.line_id),
                    "code": station.code,
                    "name": station.name
                })
            plant_data["children"].append(line_data)
        tree.append(plant_data)

    return success(data=tree)


@router.get("/lines")
async def get_lines(plant_id: str, db: AsyncSession = Depends(get_db)):
    """
    根据厂区 ID 查询线别
    """
    result = await db.execute(
        select(Line).where(Line.plant_id == plant_id)
    )
    lines = result.scalars().all()
    return success(data=[
        {"id": str(l.id), "plant_id": str(l.plant_id), "code": l.code, "name": l.name}
        for l in lines
    ])


@router.get("/stations")
async def get_stations(line_id: str, db: AsyncSession = Depends(get_db)):
    """
    根据线别 ID 查询站别
    """
    result = await db.execute(
        select(Station).where(Station.line_id == line_id)
    )
    stations = result.scalars().all()
    return success(data=[
        {"id": str(s.id), "line_id": str(s.line_id), "code": s.code, "name": s.name}
        for s in stations
    ])


@router.post("/import")
async def import_plant_dict(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """
    批量导入厂区/线别/站别数据（Excel 文件）
    
    Excel 格式（第一行为表头，从第二行开始读取）：
    | 厂区代码 | 厂区名称 | 线别代码 | 线别名称 | 站别代码 | 站别名称 |
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        return error(message="仅支持 Excel 文件（.xlsx / .xls）")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    stats = {"plant": 0, "line": 0, "station": 0}
    plant_cache = {}   # code -> Plant
    line_cache = {}    # (plant_code, line_code) -> Line

    for row in rows:
        plant_code = str(row[0]).strip() if row[0] else None
        plant_name = str(row[1]).strip() if row[1] else None
        line_code = str(row[2]).strip() if row[2] else None
        line_name = str(row[3]).strip() if row[3] else None
        station_code = str(row[4]).strip() if row[4] else None
        station_name = str(row[5]).strip() if row[5] else None

        if not plant_code or not plant_name:
            continue  # 跳过空行

        # 厂区：如果 code 不存在则创建
        if plant_code not in plant_cache:
            existing = await db.execute(select(Plant).where(Plant.code == plant_code))
            plant = existing.scalar_one_or_none()
            if not plant:
                plant = Plant(code=plant_code, name=plant_name)
                db.add(plant)
                await db.flush()
                stats["plant"] += 1
            plant_cache[plant_code] = plant
        else:
            plant = plant_cache[plant_code]

        # 线别
        if line_code and line_name:
            line_key = f"{plant_code}|{line_code}"
            if line_key not in line_cache:
                existing_line = await db.execute(
                    select(Line).where(Line.plant_id == plant.id, Line.code == line_code)
                )
                line = existing_line.scalar_one_or_none()
                if not line:
                    line = Line(plant_id=plant.id, code=line_code, name=line_name)
                    db.add(line)
                    await db.flush()
                    stats["line"] += 1
                line_cache[line_key] = line
            else:
                line = line_cache[line_key]

            # 站别
            if station_code and station_name:
                existing_station = await db.execute(
                    select(Station).where(Station.line_id == line.id, Station.code == station_code)
                )
                station = existing_station.scalar_one_or_none()
                if not station:
                    station = Station(line_id=line.id, code=station_code, name=station_name)
                    db.add(station)
                    stats["station"] += 1

    await db.commit()

    return success({
        "message": f"导入完成：新增厂区 {stats['plant']} 个，线别 {stats['line']} 个，站别 {stats['station']} 个",
        "stats": stats
    })


# ──────────── 厂区 / 线别 / 站别 新增 ────────────

@router.post("/plant")
async def create_plant(data: PlantCreate, db: AsyncSession = Depends(get_db)):
    existing = await db.execute(select(Plant).where(Plant.code == data.code))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="厂区代码已存在")
    plant = Plant(code=data.code, name=data.name)
    db.add(plant)
    await db.commit()
    await db.refresh(plant)
    return success({"id": str(plant.id), "code": plant.code, "name": plant.name})


@router.post("/line")
async def create_line(data: LineCreate, db: AsyncSession = Depends(get_db)):
    plant_result = await db.execute(select(Plant).where(Plant.id == data.plant_id))
    if not plant_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="所属厂区不存在")
    existing = await db.execute(
        select(Line).where(Line.plant_id == data.plant_id, Line.code == data.code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该厂区下线别代码已存在")
    line = Line(plant_id=data.plant_id, code=data.code, name=data.name)
    db.add(line)
    await db.commit()
    await db.refresh(line)
    return success({"id": str(line.id), "code": line.code, "name": line.name})


@router.post("/station")
async def create_station(data: StationCreate, db: AsyncSession = Depends(get_db)):
    line_result = await db.execute(select(Line).where(Line.id == data.line_id))
    if not line_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="所属线别不存在")
    existing = await db.execute(
        select(Station).where(Station.line_id == data.line_id, Station.code == data.code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该线别下站别代码已存在")
    station = Station(line_id=data.line_id, code=data.code, name=data.name)
    db.add(station)
    await db.commit()
    await db.refresh(station)
    return success({"id": str(station.id), "code": station.code, "name": station.name})


# ──────────── 厂区 修改 / 删除 ────────────

@router.put("/plant/{plant_id}")
async def update_plant(plant_id: str, data: PlantUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Plant).where(Plant.id == plant_id))
    plant = result.scalar_one_or_none()
    if not plant:
        raise HTTPException(status_code=404, detail="厂区不存在")
    if data.code is not None:
        existing = await db.execute(select(Plant).where(Plant.code == data.code, Plant.id != plant_id))
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="厂区代码已存在")
        plant.code = data.code
    if data.name is not None:
        plant.name = data.name
    await db.commit()
    return success({"id": str(plant.id), "code": plant.code, "name": plant.name})


@router.delete("/plant/{plant_id}")
async def delete_plant(plant_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Plant).where(Plant.id == plant_id))
    plant = result.scalar_one_or_none()
    if not plant:
        raise HTTPException(status_code=404, detail="厂区不存在")
    # 检查是否有巡检记录引用
    insp = await db.execute(select(Inspection).where(Inspection.plant_id == plant_id).limit(1))
    if insp.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该厂区下已有巡检记录，无法删除")
    exc = await db.execute(select(ExceptionTicket).where(ExceptionTicket.plant_id == plant_id).limit(1))
    if exc.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该厂区下已有异常工单，无法删除")
    # 删除级联的线别及其站别
    lines = (await db.execute(select(Line).where(Line.plant_id == plant_id))).scalars().all()
    for line in lines:
        await db.execute(sa_delete(Station).where(Station.line_id == line.id))
        await db.delete(line)
    await db.delete(plant)
    await db.commit()
    return success({"id": plant_id})


# ──────────── 线别 修改 / 删除 ────────────

@router.put("/line/{line_id}")
async def update_line(line_id: str, data: LineUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Line).where(Line.id == line_id))
    line = result.scalar_one_or_none()
    if not line:
        raise HTTPException(status_code=404, detail="线别不存在")
    if data.code is not None:
        existing = await db.execute(
            select(Line).where(Line.plant_id == line.plant_id, Line.code == data.code, Line.id != line_id)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="该厂区下线别代码已存在")
        line.code = data.code
    if data.name is not None:
        line.name = data.name
    await db.commit()
    return success({"id": str(line.id), "code": line.code, "name": line.name})


@router.delete("/line/{line_id}")
async def delete_line(line_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Line).where(Line.id == line_id))
    line = result.scalar_one_or_none()
    if not line:
        raise HTTPException(status_code=404, detail="线别不存在")
    insp = await db.execute(select(Inspection).where(Inspection.line_id == line_id).limit(1))
    if insp.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该线别下已有巡检记录，无法删除")
    # 删除级联的站别
    await db.execute(sa_delete(Station).where(Station.line_id == line_id))
    await db.delete(line)
    await db.commit()
    return success({"id": line_id})


# ──────────── 站别 修改 / 删除 ────────────

@router.put("/station/{station_id}")
async def update_station(station_id: str, data: StationUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Station).where(Station.id == station_id))
    station = result.scalar_one_or_none()
    if not station:
        raise HTTPException(status_code=404, detail="站别不存在")
    if data.code is not None:
        existing = await db.execute(
            select(Station).where(Station.line_id == station.line_id, Station.code == data.code, Station.id != station_id)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="该线别下站别代码已存在")
        station.code = data.code
    if data.name is not None:
        station.name = data.name
    await db.commit()
    return success({"id": str(station.id), "code": station.code, "name": station.name})


@router.delete("/station/{station_id}")
async def delete_station(station_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Station).where(Station.id == station_id))
    station = result.scalar_one_or_none()
    if not station:
        raise HTTPException(status_code=404, detail="站别不存在")
    insp = await db.execute(select(Inspection).where(Inspection.station_id == station_id).limit(1))
    if insp.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该站别下已有巡检记录，无法删除")
    await db.delete(station)
    await db.commit()
    return success({"id": station_id})