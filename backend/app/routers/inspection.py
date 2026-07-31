from fastapi import APIRouter, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import joinedload
from datetime import datetime
import uuid as uuid_module
import os
import io
import zipfile
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D
from openpyxl import load_workbook
from PIL import Image

from app.database import get_db
from app.models.inspection import Inspection, InspectionImage
from app.models.exception import ExceptionTicket
from app.models.user import SysUser
from app.models.plant_dict import Plant, Line, Station
from app.utils.response import success
from app.schemas.inspection import InspectionCreate
from app.utils.minio_client import upload_file, get_minio_client
from app.dependencies import get_current_user, get_authorized_plant_ids

router = APIRouter(prefix="/inspections", tags=["巡检记录"])


@router.post("")
async def create_inspection(
    data: InspectionCreate,
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    today = datetime.now().strftime("%Y%m%d")
    serial_no = f"INS-{today}-{uuid_module.uuid4().hex[:4].upper()}"

    inspection = Inspection(
        serial_no=serial_no,
        plant_id=data.plant_id,
        line_id=data.line_id,
        station_id=data.station_id,
        ip_address=data.ip_address,
        antivirus_status=data.antivirus_status,
        domain_status=data.domain_status,
        remark=data.remark,
        status=data.status,
        inspector_id=current_user.id
    )
    db.add(inspection)
    await db.flush()

    for img in (data.images or []):
        image = InspectionImage(
            inspection_id=inspection.id,
            file_name=img.get("name", "image.jpg"),
            storage_key=img.get("storage_key", ""),
            mime_type=img.get("mime_type", "image/jpeg")
        )
        db.add(image)

    # 异常自动建单
    if data.antivirus_status == "ABNORMAL" or data.domain_status == "NOT_JOINED":
        plant_result = await db.execute(select(Plant).where(Plant.id == data.plant_id))
        plant = plant_result.scalar_one_or_none()

        title_parts = []
        if data.antivirus_status == "ABNORMAL":
            title_parts.append("防毒软件异常")
        if data.domain_status == "NOT_JOINED":
            title_parts.append("未入域")
        title = " - ".join(title_parts)

        exception_ticket = ExceptionTicket(
            inspection_id=inspection.id,
            plant_id=data.plant_id,
            title=title,
            status="PENDING"
        )
        db.add(exception_ticket)

    await db.commit()

    return success({"id": str(inspection.id), "serial_no": serial_no})


@router.get("")
async def list_inspections(
    page: int = 1,
    page_size: int = 10,
    plant_id: str = None,
    line_id: str = None,
    station_id: str = None,
    start_time: str = None,
    end_time: str = None,
    antivirus_status: str = None,
    domain_status: str = None,
    current_user: SysUser = Depends(get_current_user),
    authorized_plant_ids: list = Depends(get_authorized_plant_ids),
    db: AsyncSession = Depends(get_db)
):
    stmt = select(Inspection).order_by(Inspection.created_at.desc())

    if not current_user.is_superadmin:
        if not authorized_plant_ids:
            return success({"list": [], "total": 0})
        stmt = stmt.where(Inspection.plant_id.in_(authorized_plant_ids))

    if plant_id:
        stmt = stmt.where(Inspection.plant_id == plant_id)
    if line_id:
        stmt = stmt.where(Inspection.line_id == line_id)
    if station_id:
        stmt = stmt.where(Inspection.station_id == station_id)
    if antivirus_status:
        stmt = stmt.where(Inspection.antivirus_status == antivirus_status)
    if domain_status:
        stmt = stmt.where(Inspection.domain_status == domain_status)
    if start_time:
        stmt = stmt.where(Inspection.inspect_time >= start_time)
    if end_time:
        stmt = stmt.where(Inspection.inspect_time <= end_time)

    # 同时查询总数
    count_stmt = select(Inspection)
    # 对 count_stmt 应用相同的过滤条件
    if not current_user.is_superadmin:
        if not authorized_plant_ids:
            count_stmt = count_stmt.where(Inspection.plant_id.in_([]))
        else:
            count_stmt = count_stmt.where(Inspection.plant_id.in_(authorized_plant_ids))
    if plant_id:
        count_stmt = count_stmt.where(Inspection.plant_id == plant_id)
    if line_id:
        count_stmt = count_stmt.where(Inspection.line_id == line_id)
    if station_id:
        count_stmt = count_stmt.where(Inspection.station_id == station_id)
    if antivirus_status:
        count_stmt = count_stmt.where(Inspection.antivirus_status == antivirus_status)
    if domain_status:
        count_stmt = count_stmt.where(Inspection.domain_status == domain_status)
    if start_time:
        count_stmt = count_stmt.where(Inspection.inspect_time >= start_time)
    if end_time:
        count_stmt = count_stmt.where(Inspection.inspect_time <= end_time)

    # 先获取总数
    total_result = await db.execute(select(
        __import__('sqlalchemy').func.count()
    ).select_from(count_stmt.subquery()))
    total = total_result.scalar() or 0

    offset = (page - 1) * page_size
    result = await db.execute(stmt.offset(offset).limit(page_size))
    inspections = result.scalars().all()

    inspection_ids = [i.id for i in inspections]
    images = []
    if inspection_ids:
        img_result = await db.execute(
            select(InspectionImage).where(InspectionImage.inspection_id.in_(inspection_ids))
        )
        images = img_result.scalars().all()

    image_map = {}
    for img in images:
        if str(img.inspection_id) not in image_map:
            image_map[str(img.inspection_id)] = []
        image_map[str(img.inspection_id)].append({
            "name": img.file_name,
            "url": f"{os.getenv('MINIO_EXTERNAL_URL', 'http://localhost:9000')}/{os.getenv('MINIO_BUCKET', 'inspection-images')}/{img.storage_key}",
            "storage_key": img.storage_key
        })

    return success({
        "list": [
            {
                "id": str(i.id),
                "serial_no": i.serial_no,
                "plant_id": str(i.plant_id),
                "line_id": str(i.line_id),
                "station_id": str(i.station_id),
                "ip_address": i.ip_address,
                "antivirus_status": i.antivirus_status,
                "domain_status": i.domain_status,
                "remark": i.remark,
                "status": i.status,
                "inspect_time": i.inspect_time.isoformat() if i.inspect_time else None,
                "images": image_map.get(str(i.id), [])
            }
            for i in inspections
        ],
        "total": total
    })


@router.post("/upload")
async def upload_inspection_image(file: UploadFile = File(...)):
    content = await file.read()
    object_name = f"inspection/{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    url = upload_file(object_name, content, file.content_type or "image/jpeg")
    return success({
        "url": url,
        "storage_key": object_name,
        "name": file.filename
    })


@router.post("/batch-import")
async def batch_import_inspections(
    file: UploadFile = File(...),
    current_user: SysUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    批量导入巡检记录（ZIP 包）
    
    ZIP 包结构：
    - records.xlsx （Excel 数据文件）
    - images/ （可选，巡检证据图片文件夹）
    
    Excel 列（第一行为表头）：
    厂区代码 | 线别代码 | 站别代码 | IP地址 | 防毒状态 | 入域状态 | 备注 | 图片数量
    
    图片命名规则：行号_序号.jpg，例如 2_1.jpg 表示第2行第1张图
    """
    if not file.filename.endswith('.zip'):
        return success(code=400, message="请上传 ZIP 压缩包")

    content = await file.read()
    zf = zipfile.ZipFile(io.BytesIO(content))
    file_list = zf.namelist()

    # 查找 Excel 文件
    excel_name = None
    for f in file_list:
        if f.endswith('.xlsx') and not f.startswith('__MACOSX'):
            excel_name = f
            break

    if not excel_name:
        return success(code=400, message="ZIP 包中未找到 .xlsx 文件")

    # 读取 Excel
    excel_data = zf.read(excel_name)
    wb = load_workbook(io.BytesIO(excel_data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    # 读取图片映射（支持多级路径和大小写）
    image_map = {}
    for f in file_list:
        if f.endswith('/') or f.startswith('__MACOSX'):
            continue
        # 检查是否是图片文件
        f_lower = f.lower()
        if not (f_lower.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'))):
            continue
        # 解析图片文件名：允许任意路径前缀，只要文件名格式为 行号_序号.jpg
        try:
            base = f.split('/')[-1].split('\\')[-1]  # 取纯文件名
            parts = base.rsplit('.', 1)[0].split('_')
            row_num = int(parts[0])
            if row_num not in image_map:
                image_map[row_num] = []
            image_map[row_num].append((f, base))
        except (ValueError, IndexError):
            continue

    stats = {"created": 0, "errors": 0, "images": 0}
    today = datetime.now().strftime("%Y%m%d")
    minio_client = get_minio_client()
    bucket = os.getenv("MINIO_BUCKET", "inspection-images")

    # 预处理：收集厂区线别站别 code -> id 映射
    # 支持格式：纯 code（如 P1）或 "code - name"（如 P1 - A厂区）
    def extract_code(val):
        raw = str(val).strip() if val else ""
        # 去除括号及内容：如 "NORMAL(正常)" → "NORMAL"
        raw = __import__('re').sub(r'\([^)]*\)', '', raw).strip()
        # 取空格或 "-" 分隔的第一部分（支持纯 code 和 "P1 - A厂区" 格式）
        parts = raw.split(None, 1)
        return parts[0] if parts else raw

    plant_codes = set()
    line_codes = set()
    station_codes = set()
    for row in rows:
        if row[0]:
            plant_codes.add(extract_code(row[0]))
        if row[1]:
            line_codes.add(extract_code(row[1]))
        if row[2]:
            station_codes.add(extract_code(row[2]))

    plant_map = {}
    if plant_codes:
        result = await db.execute(select(Plant).where(Plant.code.in_(plant_codes)))
        for p in result.scalars().all():
            plant_map[p.code] = p.id

    line_map = {}
    if line_codes:
        result = await db.execute(select(Line).where(Line.code.in_(line_codes)))
        for l in result.scalars().all():
            line_map[l.code] = l.id

    station_map = {}
    if station_codes:
        result = await db.execute(select(Station).where(Station.code.in_(station_codes)))
        for s in result.scalars().all():
            station_map[s.code] = s.id

    for row_idx, row in enumerate(rows):
        try:
            plant_code = extract_code(row[0]) if row[0] else None
            line_code = extract_code(row[1]) if row[1] else None
            station_code = extract_code(row[2]) if row[2] else None
            ip_address = str(row[3]).strip() if row[3] else None
            antivirus_status = str(row[4]).strip() if row[4] else "NORMAL"
            domain_status = str(row[5]).strip() if row[5] else "JOINED"
            remark = str(row[6]).strip() if row[6] else None
            _img_count = row[7]  # 图片数量（仅作参考）

            if not plant_code or not line_code or not station_code or not ip_address:
                stats["errors"] += 1
                continue

            plant_id = plant_map.get(plant_code)
            line_id = line_map.get(line_code)
            station_id = station_map.get(station_code)

            if not plant_id or not line_id or not station_id:
                stats["errors"] += 1
                continue

            # IP 地址校验
            ip_parts = ip_address.split('.')
            if len(ip_parts) != 4:
                stats["errors"] += 1
                continue

            # 状态标准化（支持带括号格式如 NORMAL(正常)、已入域 等）
            antivirus_status = extract_code(antivirus_status)  # 去除括号
            if antivirus_status not in ("NORMAL", "ABNORMAL", "NOT_INSTALLED"):
                antivirus_status_map = {"正常": "NORMAL", "异常": "ABNORMAL", "未安装": "NOT_INSTALLED"}
                antivirus_status = antivirus_status_map.get(antivirus_status, "NORMAL")

            domain_status = extract_code(domain_status)  # 去除括号
            if domain_status not in ("JOINED", "NOT_JOINED", "NOT_APPLICABLE"):
                domain_status_map = {"已入域": "JOINED", "未入域": "NOT_JOINED", "不适用": "NOT_APPLICABLE"}
                domain_status = domain_status_map.get(domain_status, "JOINED")

            serial_no = f"INS-{today}-{uuid_module.uuid4().hex[:4].upper()}"

            inspection = Inspection(
                serial_no=serial_no,
                plant_id=plant_id,
                line_id=line_id,
                station_id=station_id,
                ip_address=ip_address,
                antivirus_status=antivirus_status,
                domain_status=domain_status,
                remark=remark,
                status="SUBMITTED",
                inspector_id=current_user.id
            )
            db.add(inspection)
            await db.flush()

            # 处理巡检证据图片
            excel_row_num = row_idx + 2  # Excel 行号（第1行是表头）
            if excel_row_num in image_map:
                for img_path, img_name in image_map[excel_row_num]:
                    try:
                        img_data = zf.read(img_path)
                        object_name = f"inspection/{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid_module.uuid4().hex[:6]}_{img_name}"

                        minio_client.put_object(
                            bucket,
                            object_name,
                            data=io.BytesIO(img_data),
                            length=len(img_data),
                            content_type="image/jpeg"
                        )

                        image = InspectionImage(
                            inspection_id=inspection.id,
                            file_name=img_name,
                            storage_key=object_name,
                            mime_type="image/jpeg"
                        )
                        db.add(image)
                        stats["images"] += 1
                    except Exception:
                        pass  # 单张图片失败不中断

            # 异常自动建单
            if antivirus_status == "ABNORMAL" or domain_status == "NOT_JOINED":
                title_parts = []
                if antivirus_status == "ABNORMAL":
                    title_parts.append("防毒软件异常")
                if domain_status == "NOT_JOINED":
                    title_parts.append("未入域")
                title = " - ".join(title_parts)

                exception_ticket = ExceptionTicket(
                    inspection_id=inspection.id,
                    plant_id=plant_id,
                    title=title,
                    status="PENDING"
                )
                db.add(exception_ticket)

            stats["created"] += 1
        except Exception:
            stats["errors"] += 1

    await db.commit()

    return success({
        "message": f"导入完成：成功 {stats['created']} 条，失败 {stats['errors']} 条，图片 {stats['images']} 张",
        "stats": stats
    })


# 状态中文映射
STATUS_LABEL_MAP = {
    "NORMAL": "正常",
    "ABNORMAL": "异常",
    "NOT_INSTALLED": "未安装",
    "JOINED": "已入域",
    "NOT_JOINED": "未入域",
    "NOT_APPLICABLE": "不适用",
}


def _build_inspection_query(
    stmt,
    plant_id: str = None,
    line_id: str = None,
    station_id: str = None,
    start_time: str = None,
    end_time: str = None,
    antivirus_status: str = None,
    domain_status: str = None,
):
    """构建巡检记录查询的通用筛选条件"""
    if plant_id:
        stmt = stmt.where(Inspection.plant_id == plant_id)
    if line_id:
        stmt = stmt.where(Inspection.line_id == line_id)
    if station_id:
        stmt = stmt.where(Inspection.station_id == station_id)
    if antivirus_status:
        stmt = stmt.where(Inspection.antivirus_status == antivirus_status)
    if domain_status:
        stmt = stmt.where(Inspection.domain_status == domain_status)
    if start_time:
        stmt = stmt.where(Inspection.inspect_time >= start_time)
    if end_time:
        stmt = stmt.where(Inspection.inspect_time <= end_time)
    return stmt


@router.get("/export")
async def export_inspections(
    plant_id: str = Query(None),
    line_id: str = Query(None),
    station_id: str = Query(None),
    start_time: str = Query(None),
    end_time: str = Query(None),
    antivirus_status: str = Query(None),
    domain_status: str = Query(None),
    current_user: SysUser = Depends(get_current_user),
    authorized_plant_ids: list = Depends(get_authorized_plant_ids),
    db: AsyncSession = Depends(get_db),
):
    """导出巡检记录为 Excel 文件"""
    # 联表查询：巡检记录 + 厂区/线别/站别 + 巡检人
    stmt = (
        select(Inspection)
        .options(
            joinedload(Inspection.images),
        )
        .order_by(Inspection.created_at.desc())
    )

    # 厂区权限过滤
    if not current_user.is_superadmin:
        if not authorized_plant_ids:
            stmt = stmt.where(Inspection.plant_id.in_([]))
        else:
            stmt = stmt.where(Inspection.plant_id.in_(authorized_plant_ids))

    stmt = _build_inspection_query(
        stmt,
        plant_id=plant_id,
        line_id=line_id,
        station_id=station_id,
        start_time=start_time,
        end_time=end_time,
        antivirus_status=antivirus_status,
        domain_status=domain_status,
    )

    result = await db.execute(stmt)
    inspections = result.unique().scalars().all()

    # 收集关联 ID 并批量查询名称
    plant_ids = list({i.plant_id for i in inspections})
    line_ids = list({i.line_id for i in inspections})
    station_ids = list({i.station_id for i in inspections})
    inspector_ids = list({i.inspector_id for i in inspections})

    plant_name_map = {}
    if plant_ids:
        plant_result = await db.execute(select(Plant).where(Plant.id.in_(plant_ids)))
        for p in plant_result.scalars().all():
            plant_name_map[str(p.id)] = f"{p.code} - {p.name}"

    line_name_map = {}
    if line_ids:
        line_result = await db.execute(select(Line).where(Line.id.in_(line_ids)))
        for l in line_result.scalars().all():
            line_name_map[str(l.id)] = f"{l.code} - {l.name}"

    station_name_map = {}
    if station_ids:
        station_result = await db.execute(select(Station).where(Station.id.in_(station_ids)))
        for s in station_result.scalars().all():
            station_name_map[str(s.id)] = f"{s.code} - {s.name}"

    inspector_name_map = {}
    if inspector_ids:
        user_result = await db.execute(select(SysUser).where(SysUser.id.in_(inspector_ids)))
        for u in user_result.scalars().all():
            inspector_name_map[str(u.id)] = u.real_name or u.username

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "巡检记录"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 表头
    headers = ["巡检单号", "厂区", "线别", "站别", "IP地址", "防毒软件状态", "入域状态", "备注", "巡检时间", "巡检人", "巡检证据"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 初始化 MinIO 客户端
    minio_client = get_minio_client()
    bucket_name = os.getenv("MINIO_BUCKET", "inspection-images")

    # 图片嵌入配置
    IMG_TARGET_HEIGHT = 90
    IMAGES_PER_ROW = 3
    IMG_GAP_PX = 4
    IMG_EMBED_WIDTH = (IMG_TARGET_HEIGHT * IMAGES_PER_ROW + IMG_GAP_PX * (IMAGES_PER_ROW + 1)) / 7.0

    # 数据行
    for row_idx, inspection in enumerate(inspections, 2):
        values = [
            inspection.serial_no,
            plant_name_map.get(str(inspection.plant_id), ""),
            line_name_map.get(str(inspection.line_id), ""),
            station_name_map.get(str(inspection.station_id), ""),
            inspection.ip_address,
            STATUS_LABEL_MAP.get(inspection.antivirus_status, inspection.antivirus_status),
            STATUS_LABEL_MAP.get(inspection.domain_status, inspection.domain_status),
            inspection.remark or "",
            inspection.inspect_time.strftime("%Y-%m-%d %H:%M:%S") if inspection.inspect_time else "",
            inspector_name_map.get(str(inspection.inspector_id), ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = thin_border
            if col_idx in (8,):
                cell.alignment = cell_alignment_left
            else:
                cell.alignment = cell_alignment

        evidence_cell = ws.cell(row=row_idx, column=11)
        evidence_cell.border = thin_border
        evidence_cell.alignment = Alignment(horizontal="center", vertical="top")

        images = inspection.images
        if images:
            img_rows_count = (len(images) + IMAGES_PER_ROW - 1) // IMAGES_PER_ROW
            row_height_px = img_rows_count * (IMG_TARGET_HEIGHT + IMG_GAP_PX) + IMG_GAP_PX
            ws.row_dimensions[row_idx].height = row_height_px * 0.75

            for img_idx, inspection_image in enumerate(images):
                try:
                    img_obj = minio_client.get_object(bucket_name, inspection_image.storage_key)
                    img_bytes = img_obj.read()
                    img_obj.close()
                    img_obj.release_conn()

                    pil_img = Image.open(io.BytesIO(img_bytes))
                    pil_img = pil_img.convert("RGB")
                    w, h = pil_img.size
                    new_w = int(w * IMG_TARGET_HEIGHT / h)
                    pil_img = pil_img.resize((new_w, IMG_TARGET_HEIGHT), Image.LANCZOS)

                    img_stream = io.BytesIO()
                    pil_img.save(img_stream, format="JPEG", quality=85)
                    img_stream.seek(0)

                    img_col = img_idx % IMAGES_PER_ROW
                    img_row = img_idx // IMAGES_PER_ROW
                    offset_x = IMG_GAP_PX + img_col * (IMG_TARGET_HEIGHT + IMG_GAP_PX)
                    offset_y = IMG_GAP_PX + img_row * (IMG_TARGET_HEIGHT + IMG_GAP_PX)

                    xl_img = XLImage(img_stream)
                    xl_img.width = new_w
                    xl_img.height = IMG_TARGET_HEIGHT

                    col_offset = int(offset_x * 9525)
                    row_offset = int(offset_y * 9525)

                    xl_img.anchor = OneCellAnchor(
                        _from=AnchorMarker(
                            col=10,
                            colOff=col_offset,
                            row=row_idx - 1,
                            rowOff=row_offset,
                        ),
                        ext=XDRPositiveSize2D(new_w * 9525, IMG_TARGET_HEIGHT * 9525),
                    )

                    ws.add_image(xl_img)
                except Exception:
                    evidence_cell.value = (evidence_cell.value or "") + f"[图片加载失败] "
        else:
            evidence_cell.value = "无"
            ws.row_dimensions[row_idx].height = 30

    column_widths = [20, 22, 22, 22, 16, 16, 12, 30, 22, 14, IMG_EMBED_WIDTH]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"巡检记录_{datetime.now().strftime('%Y%m%d')}.xlsx"
    encoded_filename = quote(filename, safe='')

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )